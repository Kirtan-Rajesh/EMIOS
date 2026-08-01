"""Shared types for the Document Discovery extractors (app/services_v1/extraction/).

Every extractor (SchemaExtractor, StoredProcExtractor, ApiSpecExtractor,
EtlMappingExtractor, MetadataExtractor, LlmPromptExtractor) takes one uploaded
file's raw bytes and returns an ExtractionResult - a partial, per-document
contribution of ServiceNode/DependencyEdge objects in the exact shape the rest
of the platform (GraphService, the 12-agent pipeline, simulation, wave
planning) already consumes. DocumentExtractionService is what fans a batch of
uploaded documents out to the right extractor and merges their results.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Dict, List

from app.models.schemas import DependencyEdge, ServiceNode

# Shared by SchemaExtractor and StoredProcExtractor so a schema dump and a
# stored-proc dump describing the same system converge on the same node id
# (e.g. "payments.transactions" and "payments.sp_settle" both -> "payments_db")
# without relying on the fuzzy entity-resolution pass to reconcile them.
UNQUALIFIED_SCHEMA = "default"


def safe_node_id(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return slug or "schema"


@dataclass
class ExtractionResult:
    source_filename: str
    extractor: str
    nodes: List[ServiceNode] = field(default_factory=list)
    edges: List[DependencyEdge] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    # 1.0 for deterministic parsers (schema/API-spec/ETL-mapping/stored-proc);
    # lower and caller-facing for the LLM-based BRD/HLD extractor, whose output
    # is a best-effort reading of prose rather than a parsed structure.
    confidence: float = 1.0


class BaseExtractor:
    """Common interface every extractor implements. `can_handle` is a cheap,
    best-effort content sniff (filename + a text preview, plus the raw bytes
    for binary formats like XLSX where a decoded text preview is garbage)
    used by DocumentExtractionService's dispatcher - it is not authoritative
    on its own; extractors should still degrade gracefully (return an
    ExtractionResult with warnings, never raise) if handed content that turns
    out not to match."""

    name: str = "BaseExtractor"

    def can_handle(self, filename: str, text_preview: str, content: bytes = b"") -> bool:
        raise NotImplementedError

    def extract(self, filename: str, content: bytes, text_preview: str) -> ExtractionResult:
        raise NotImplementedError


def sniff_header_line(filename: str, content: bytes) -> str:
    """Returns the lowercased header row as one comma-joined line, for the
    cheap substring-based sniffing MetadataExtractor/EtlMappingExtractor use
    in can_handle(). CSV's header is just its first text line; XLSX is a
    binary zip container, so a decoded-as-text preview is garbage - this
    parses the first row of the first worksheet instead. Either way, callers
    get one uniform string to run the same substring checks against,
    regardless of which tabular format a team happened to export."""
    if filename.lower().endswith(".xlsx"):
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            first_row = next(workbook.worksheets[0].iter_rows(values_only=True), ())
            return ",".join("" if c is None else str(c) for c in first_row).lower()
        except Exception:
            return ""
    text_preview = content[:4000].decode("utf-8", errors="ignore")
    lines = text_preview.splitlines()
    return lines[0].lower() if lines else ""


def read_tabular_rows(filename: str, content: bytes) -> List[Dict[str, str]]:
    """Reads a CSV or XLSX file into a list of header -> value dict rows (the
    same shape csv.DictReader produces), so MetadataExtractor/EtlMappingExtractor
    can share one row-reading path regardless of tabular format. XLSX reads
    only the first (active) worksheet - these extractors are purpose-built for
    single-table catalogs/mappings, not multi-sheet workbooks; a multi-sheet
    export should be split into one file per sheet."""
    if filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows_iter = workbook.worksheets[0].iter_rows(values_only=True)
        try:
            header = ["" if h is None else str(h).strip() for h in next(rows_iter)]
        except StopIteration:
            return []
        rows: List[Dict[str, str]] = []
        for row in rows_iter:
            if all(cell is None for cell in row):
                continue
            rows.append(
                {
                    header[i]: ("" if i >= len(row) or row[i] is None else str(row[i]))
                    for i in range(len(header))
                    if header[i]
                }
            )
        return rows

    text = content.decode("utf-8", errors="replace")
    return [dict(row) for row in csv.DictReader(io.StringIO(text))]
